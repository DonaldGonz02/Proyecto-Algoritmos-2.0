import sys
from os import path
from pathlib import Path
import smtplib
import ssl
import ctypes
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.utils import formatdate
from email import encoders

# Librerías a instalar
from PyQt5 import QtCore
from PyQt5 import QtGui
from PyQt5 import QtWidgets
import xlsxwriter
import openpyxl


# Revisa que haya un archivo excel y de no haber, crea uno
archivo_excel = "Inventario.xlsx"
if not path.isfile(archivo_excel):
    workbook = xlsxwriter.Workbook(archivo_excel)
    worksheet_productos = workbook.add_worksheet("productos")
    worksheet_productos.write_row("A1", ["nombre", "precio", "existencia"])

    worksheet_clientes = workbook.add_worksheet("cliente")
    worksheet_clientes.write_row("A1", ["nombre", "nit", "dirección"])

    worksheet_pedidos = workbook.add_worksheet("pedido")
    worksheet_pedidos.write_row("A1", ["nombre cliente", "nombre producto", "cantidad producto", "valor pedido"])

    workbook.close()


# Base de datos, lee y guarda los datos a un archivo excel
class BaseDatos:
    datos = {
        "productos": {
            "títulos": ["nombre", "precio", "existencia"],
            "datos": []
        },
        "cliente": {
            "títulos": ["nombre", "nit", "dirección"],
            "datos": []
        },
        "pedido": {
            "títulos": ["nombre cliente", "nombre producto", "cantidad producto", "valor pedido"],
            "datos": []
        }
    }

    def __init__(self):
        self.archivo = Path(archivo_excel)
        self.leer_excel()

    def leer_excel(self):
        wb_obj = openpyxl.load_workbook(self.archivo)
        for hoja in wb_obj.sheetnames:
            sheet_wb = wb_obj[hoja]
            data = []
            for fila in sheet_wb.iter_rows(max_row=sheet_wb.max_row):
                temp = []
                for cell in fila:
                    temp.append(cell.value)
                data.append(temp)
            self.datos[hoja]["datos"] = data[1:]

    def guardar_excel(self, route=None):
        workbook = xlsxwriter.Workbook(archivo_excel if route is None else route)
        for pagina in self.datos:
            worksheet = workbook.add_worksheet(pagina)
            worksheet.write_row("A1", self.datos[pagina]["títulos"])
            for i, fila in enumerate(self.datos[pagina]["datos"]):
                worksheet.write_row(f"A{i+2}", fila)
        workbook.close()

    def borrar_item(self, index: int, table: str):
        self.datos[table]["datos"].pop(index)
        self.guardar_excel()

    def crear_producto(self, name: str, quantity: int, price: float):
        self.datos["productos"]["datos"].append([name, price, quantity])
        self.guardar_excel()

    def editar_producto(self, index: int, name: str, quantity: int, price: float):
        self.datos["productos"]["datos"][index] = [name, price, quantity]
        self.guardar_excel()

    def crear_cliente(self, name: str, nit: str, address: str):
        self.datos["cliente"]["datos"].append([name, nit, address])
        self.guardar_excel()

    def editar_cliente(self, index: int, name: str, nit: str, address: str):
        self.datos["cliente"]["datos"][index] = [name, nit, address]
        self.guardar_excel()

    def crear_pedido(self, name: str, product: str, quantity: int, price: float):
        self.datos["pedido"]["datos"].append([name, product, quantity, price])
        self.guardar_excel()

    def actualizar_existencia(self, index: int, minus=0):
        self.datos["productos"]["datos"][index][2] = self.datos["productos"]["datos"][index][2] - minus
        self.guardar_excel()


# Animación de desliz entre pestañas
class PySlidingStackedWidget(QtWidgets.QStackedWidget):
    def __init__(self, parent=None):
        super(PySlidingStackedWidget, self).__init__(parent)

        self.m_direction = QtCore.Qt.Horizontal
        self.m_speed = 500
        self.m_animationtype = QtCore.QEasingCurve.InOutCubic
        self.m_now = 0
        self.m_next = 0
        self.m_wrap = False
        self.m_pnow = QtCore.QPoint(0, 0)
        self.m_active = False

    def setDirection(self, direction):
        self.m_direction = direction

    def setSpeed(self, speed):
        self.m_speed = speed

    def setAnimation(self, animationtype):
        self.m_animationtype = animationtype

    def setWrap(self, wrap):
        self.m_wrap = wrap

    @QtCore.pyqtSlot()
    def slideInPrev(self):
        now = self.currentIndex()
        if self.m_wrap or now > 0:
            self.slideInIdx(now - 1)

    @QtCore.pyqtSlot()
    def slideInNext(self):
        now = self.currentIndex()
        if self.m_wrap or now < (self.count() - 1):
            self.slideInIdx(now + 1)

    def slideInIdx(self, idx):
        if idx > (self.count() - 1):
            idx = idx % self.count()
        elif idx < 0:
            idx = (idx + self.count()) % self.count()
        self.slideInWgt(self.widget(idx))

    def slideInWgt(self, newwidget):
        if self.m_active:
            return

        self.m_active = True

        _now = self.currentIndex()
        _next = self.indexOf(newwidget)

        if _now == _next:
            self.m_active = False
            return

        offsetx, offsety = self.frameRect().width(), self.frameRect().height()
        self.widget(_next).setGeometry(self.frameRect())

        if not self.m_direction == QtCore.Qt.Horizontal:
            if _now < _next:
                offsetx, offsety = 0, -offsety
            else:
                offsetx = 0
        else:
            if _now < _next:
                offsetx, offsety = -offsetx, 0
            else:
                offsety = 0

        pnext = self.widget(_next).pos()
        pnow = self.widget(_now).pos()
        self.m_pnow = pnow

        offset = QtCore.QPoint(offsetx, offsety)
        self.widget(_next).move(pnext - offset)
        self.widget(_next).show()
        self.widget(_next).raise_()

        # noinspection PyArgumentList
        anim_group = QtCore.QParallelAnimationGroup(
            self, finished=self.animationDoneSlot
        )

        for index, start, end in zip(
            (_now, _next), (pnow, pnext - offset), (pnow + offset, pnext)
        ):
            # noinspection PyArgumentList
            animation = QtCore.QPropertyAnimation(
                self.widget(index),
                b"pos",
                duration=self.m_speed,
                easingCurve=self.m_animationtype,
                startValue=start,
                endValue=end,
            )
            anim_group.addAnimation(animation)

        self.m_next = _next
        self.m_now = _now
        self.m_active = True
        anim_group.start(QtCore.QAbstractAnimation.DeleteWhenStopped)

    @QtCore.pyqtSlot()
    def animationDoneSlot(self):
        self.setCurrentIndex(self.m_next)
        self.widget(self.m_now).hide()
        self.widget(self.m_now).move(self.m_pnow)
        self.m_active = False


# Animación de borrado de items
class AnimatedFrame(QtWidgets.QFrame):
    def __init__(self, custom_size=200, *args, **kwargs):
        super(AnimatedFrame, self).__init__(*args, **kwargs)
        self.setMaximumHeight(0)

        self.animation = QtCore.QPropertyAnimation(self, b"maximumHeight")
        self.animation.setDuration(200)
        self.animation.setStartValue(0)
        self.animation.setEndValue(custom_size)

        self.__post_init__()

    def remove(self, action=True):
        opacity_effect = QtWidgets.QGraphicsOpacityEffect(self)
        self.setGraphicsEffect(opacity_effect)

        self.setStyleSheet("background: rgba(255, 100, 100, 150);")

        self.anim_1 = QtCore.QPropertyAnimation(opacity_effect, b"opacity")
        self.anim_1.setDuration(150)
        self.anim_1.setStartValue(1)
        self.anim_1.setEndValue(0)
        self.anim_1.setEasingCurve(QtCore.QEasingCurve.Linear)

        self.anim_2 = QtCore.QPropertyAnimation(self, b"maximumWidth")
        self.anim_2.setDuration(200)
        self.anim_2.setStartValue(self.width())
        self.anim_2.setEndValue(0)
        self.anim_2.setEasingCurve(QtCore.QEasingCurve.InOutCubic)

        self.anim_3 = QtCore.QPropertyAnimation(self, b"maximumHeight")
        self.anim_3.setDuration(300)
        self.anim_3.setStartValue(self.height())
        self.anim_3.setEndValue(0)
        self.anim_3.setEasingCurve(QtCore.QEasingCurve.InOutCubic)

        self.animation_group = QtCore.QParallelAnimationGroup(self)
        self.animation_group.addAnimation(self.anim_1)
        self.animation_group.addAnimation(self.anim_2)
        self.animation_group.addAnimation(self.anim_3)
        if action:
            self.animation_group.finished.connect(lambda: self.setParent(None))
        else:
            self.animation_group.finished.connect(self.hide)

        self.animation_group.start()

    def __post_init__(self):
        QtCore.QTimer.singleShot(100, self.animation.start)


# Ventana de creación y edición de productos
class CreacionEdicionProductos(QtWidgets.QDialog):
    cerrado = QtCore.pyqtSignal()
    
    # creación de interfaz
    def __init__(self, basededatos, index: int = None):
        super(CreacionEdicionProductos, self).__init__()
        self.setWindowFlags(QtCore.Qt.WindowFlags(QtCore.Qt.FramelessWindowHint))
        self.setAttribute(QtCore.Qt.WA_TranslucentBackground)
        self.layout = QtWidgets.QVBoxLayout()
        self.layout.setContentsMargins(0, 0, 0, 0)

        main_frame = QtWidgets.QFrame()
        main_frame.setObjectName("main_frame")
        main_frame_layout = QtWidgets.QVBoxLayout()
        main_frame_layout.setContentsMargins(0, 0, 0, 0)

        self.setStyleSheet("""
        *{
            border-radius: 10px;
            font-family:century-gothic;
            font-size:24px;
        }
        * QFrame {
            background-color: rgba(30, 50, 120, 255);
        }
        * QLabel{
            background-color: rgba(255, 255, 255, 0);
            color: rgba(255, 255, 255, 255);
            font-size: 14px;
        }
        * QPushButton
        {
            background: rgba(255, 200, 200, 80);
            color: white;
        }
        * QPushButton:hover
        {	
            background: rgba(255, 200, 200, 100);
            color: white;
        }
        * QPushButton:disabled {
            background-color: white;
            color: black;
        }
        * QScrollBar:horizontal {
            border: none;
            background: rgb(52, 59, 72);
            height: 8px;
            margin: 0px 21px 0 21px;
            border-radius: 0px;
        }
        * QScrollBar::handle:horizontal {
            background: rgb(189, 147, 249);
            min-width: 25px;
            border-radius: 4px
        }
        * QScrollBar::add-line:horizontal {
            border: none;
            background: rgb(55, 63, 77);
            width: 20px;
            border-top-right-radius: 4px;
            border-bottom-right-radius: 4px;
            subcontrol-position: right;
            subcontrol-origin: margin;
        }
        * QScrollBar::sub-line:horizontal {
            border: none;
            background: rgb(55, 63, 77);
            width: 20px;
            border-top-left-radius: 4px;
            border-bottom-left-radius: 4px;
            subcontrol-position: left;
            subcontrol-origin: margin;
        }
        * QScrollBar::up-arrow:horizontal, QScrollBar::down-arrow:horizontal
        {
             background: none;
        }
        * QScrollBar::add-page:horizontal, QScrollBar::sub-page:horizontal
        {
             background: none;
        }
        * QScrollBar:vertical {
            border: none;
            background: rgb(52, 59, 72);
            width: 8px;
            margin: 21px 0 21px 0;
            border-radius: 0px;
        }
        * QScrollBar::handle:vertical {	
           background: rgb(189, 147, 249);
           min-height: 25px;
           border-radius: 4px
        }
        * QScrollBar::add-line:vertical {
            border: none;
           background: rgb(55, 63, 77);
            height: 20px;
           border-bottom-left-radius: 4px;
           border-bottom-right-radius: 4px;
            subcontrol-position: bottom;
            subcontrol-origin: margin;
        }
        * QScrollBar::sub-line:vertical {
           border: none;
           background: rgb(55, 63, 77);
            height: 20px;
           border-top-left-radius: 4px;
           border-top-right-radius: 4px;
            subcontrol-position: top;
            subcontrol-origin: margin;
        }
        * QScrollBar::up-arrow:vertical, QScrollBar::down-arrow:vertical {
            background: none;
        }
        * QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical {
            background: none;
        }
        """)

        self.basededatos = basededatos
        self.index = index

        title = QtWidgets.QLabel("Creación de Productos" if self.index is None else "Edición de Producto")
        title.setMargin(6)
        title.setAlignment(QtCore.Qt.AlignCenter)
        title.setStyleSheet("font-size:24px; background: rgba(0, 0, 0, 40)")
        main_frame_layout.addWidget(title)

        layout_items = QtWidgets.QHBoxLayout()
        layout_items.setContentsMargins(9, 0, 9, 0)
        layout_items.setSpacing(14)

        name_layout = QtWidgets.QVBoxLayout()
        name_layout.setSpacing(6)
        self.name = QtWidgets.QLineEdit("" if self.index is None else str(self.basededatos.datos["productos"]["datos"][index][0]))
        self.name.textChanged.connect(self.Revisor)
        self.name.setTextMargins(6, 6, 6, 6)
        name_layout.addWidget(QtWidgets.QLabel("Nombre del Producto:"))
        name_layout.addWidget(self.name)

        layout_items.addLayout(name_layout)

        quantity_layout = QtWidgets.QVBoxLayout()
        quantity_layout.setSpacing(6)
        self.quantity = QtWidgets.QSpinBox()
        self.quantity.lineEdit().setTextMargins(6, 6, 6, 6)
        self.quantity.setAlignment(QtCore.Qt.AlignHCenter | QtCore.Qt.AlignVCenter)
        self.quantity.setButtonSymbols(QtWidgets.QAbstractSpinBox().NoButtons)
        self.quantity.setMinimum(0)
        self.quantity.setMaximum(999999999)
        self.quantity.clear()
        if self.index is not None:
            self.quantity.setValue(int(self.basededatos.datos["productos"]["datos"][index][2]))
        quantity_layout.addWidget(QtWidgets.QLabel("Cantidad de Items:"))
        quantity_layout.addWidget(self.quantity)

        layout_items.addLayout(quantity_layout)

        price_layout = QtWidgets.QVBoxLayout()
        price_layout.setSpacing(6)
        self.price = QtWidgets.QDoubleSpinBox()
        self.price.lineEdit().setTextMargins(6, 6, 6, 6)
        self.price.setAlignment(QtCore.Qt.AlignHCenter | QtCore.Qt.AlignVCenter)
        self.price.setButtonSymbols(QtWidgets.QAbstractSpinBox().NoButtons)
        self.price.setMinimum(0.00)
        self.price.setMaximum(999999999.99)
        self.price.clear()
        if self.index is not None:
            self.price.setValue(float(self.basededatos.datos["productos"]["datos"][index][1]))
        price_layout.addWidget(QtWidgets.QLabel("Precio Producto:"))
        price_layout.addWidget(self.price)

        layout_items.addLayout(price_layout)

        layout_buttons = QtWidgets.QHBoxLayout()
        layout_buttons.setContentsMargins(9, 0, 9, 9)

        cancel = QtWidgets.QPushButton("Cancelar ")
        cancel.setStyleSheet("""
        QPushButton
        {
            background: rgba(255, 0, 0, 100);
            color: white;
        }
        QPushButton:hover
        {	
            background: rgba(255, 0, 0, 120);
            color: white;
        }
        """)
        cancel.setIconSize(QtCore.QSize(31, 31))
        cancel.setIcon(QtGui.QIcon("icons/round_cancel_white_48dp.png"))
        cancel.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        cancel.clicked.connect(self.close)
        layout_buttons.addWidget(cancel)

        layout_buttons.addSpacerItem(
            QtWidgets.QSpacerItem(
                0, 10, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding
            )
        )

        self.save = QtWidgets.QPushButton("Guardar ")
        self.save.setStyleSheet("""
                QPushButton
                {
                    background: rgba(100, 255, 100, 80);
                    color: white;
                }
                QPushButton:hover
                {	
                    background: rgba(100, 255, 100, 100);
                    color: white;
                }
                """)
        self.save.setIconSize(QtCore.QSize(31, 31))
        self.save.setIcon(QtGui.QIcon("icons/round_check_circle_white_48dp.png"))
        self.save.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.save.clicked.connect(self.guardar_producto)
        layout_buttons.addWidget(self.save)

        main_frame_layout.addLayout(layout_items)
        main_frame_layout.addLayout(layout_buttons)
        main_frame.setLayout(main_frame_layout)
        self.layout.addWidget(main_frame)
        self.setLayout(self.layout)
        self.Revisor()
    
    # Revisor de datos para guardar
    def Revisor(self):
        revisor = True

        if self.name.text() == "":
            revisor = False

        if revisor:
            self.save.show()
        else:
            self.save.hide()
    
    # Guardar el producto
    def guardar_producto(self):
        if self.index is None:
            self.basededatos.crear_producto(name=self.name.text(), quantity=self.quantity.value(), price=self.price.value())
        else:
            self.basededatos.editar_producto(index=self.index, name=self.name.text(), quantity=self.quantity.value(), price=self.price.value())
        self.close()
        self.cerrado.emit()


# Ventana de creación y edición de clientes
class CreacionEdicionClientes(QtWidgets.QDialog):
    cerrado = QtCore.pyqtSignal()

    def __init__(self, basededatos, index: int = None):
        super(CreacionEdicionClientes, self).__init__()
        self.setWindowFlags(QtCore.Qt.WindowFlags(QtCore.Qt.FramelessWindowHint))
        self.setAttribute(QtCore.Qt.WA_TranslucentBackground)
        self.layout = QtWidgets.QVBoxLayout()
        self.layout.setContentsMargins(0, 0, 0, 0)

        main_frame = QtWidgets.QFrame()
        main_frame.setObjectName("main_frame")
        main_frame_layout = QtWidgets.QVBoxLayout()
        main_frame_layout.setContentsMargins(0, 0, 0, 0)

        self.setStyleSheet("""
        *{
            border-radius: 10px;
            font-family:century-gothic;
            font-size:24px;
        }
        * QFrame {
            background-color: rgba(30, 50, 120, 255);
        }
        * QLabel{
            background-color: rgba(255, 255, 255, 0);
            color: rgba(255, 255, 255, 255);
            font-size: 14px;
        }
        * QPushButton
        {
            background: rgba(255, 200, 200, 80);
            color: white;
        }
        * QPushButton:hover
        {	
            background: rgba(255, 200, 200, 100);
            color: white;
        }
        * QPushButton:disabled {
            background-color: white;
            color: black;
        }
        * QScrollBar:horizontal {
            border: none;
            background: rgb(52, 59, 72);
            height: 8px;
            margin: 0px 21px 0 21px;
            border-radius: 0px;
        }
        * QScrollBar::handle:horizontal {
            background: rgb(189, 147, 249);
            min-width: 25px;
            border-radius: 4px
        }
        * QScrollBar::add-line:horizontal {
            border: none;
            background: rgb(55, 63, 77);
            width: 20px;
            border-top-right-radius: 4px;
            border-bottom-right-radius: 4px;
            subcontrol-position: right;
            subcontrol-origin: margin;
        }
        * QScrollBar::sub-line:horizontal {
            border: none;
            background: rgb(55, 63, 77);
            width: 20px;
            border-top-left-radius: 4px;
            border-bottom-left-radius: 4px;
            subcontrol-position: left;
            subcontrol-origin: margin;
        }
        * QScrollBar::up-arrow:horizontal, QScrollBar::down-arrow:horizontal
        {
             background: none;
        }
        * QScrollBar::add-page:horizontal, QScrollBar::sub-page:horizontal
        {
             background: none;
        }
        * QScrollBar:vertical {
            border: none;
            background: rgb(52, 59, 72);
            width: 8px;
            margin: 21px 0 21px 0;
            border-radius: 0px;
        }
        * QScrollBar::handle:vertical {	
           background: rgb(189, 147, 249);
           min-height: 25px;
           border-radius: 4px
        }
        * QScrollBar::add-line:vertical {
            border: none;
           background: rgb(55, 63, 77);
            height: 20px;
           border-bottom-left-radius: 4px;
           border-bottom-right-radius: 4px;
            subcontrol-position: bottom;
            subcontrol-origin: margin;
        }
        * QScrollBar::sub-line:vertical {
           border: none;
           background: rgb(55, 63, 77);
            height: 20px;
           border-top-left-radius: 4px;
           border-top-right-radius: 4px;
            subcontrol-position: top;
            subcontrol-origin: margin;
        }
        * QScrollBar::up-arrow:vertical, QScrollBar::down-arrow:vertical {
            background: none;
        }
        * QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical {
            background: none;
        }
        """)

        self.basededatos = basededatos
        self.index = index

        title = QtWidgets.QLabel("Creación de Cliente" if self.index is None else "Edición de Cliente")
        title.setMargin(6)
        title.setAlignment(QtCore.Qt.AlignCenter)
        title.setStyleSheet("font-size:24px; background: rgba(0, 0, 0, 40)")
        main_frame_layout.addWidget(title)

        layout_items = QtWidgets.QHBoxLayout()
        layout_items.setContentsMargins(9, 0, 9, 0)
        layout_items.setSpacing(14)

        name_layout = QtWidgets.QVBoxLayout()
        name_layout.setSpacing(6)
        self.name = QtWidgets.QLineEdit("" if self.index is None else str(self.basededatos.datos["cliente"]["datos"][index][0]))
        self.name.textChanged.connect(self.Revisor)
        self.name.setTextMargins(6, 6, 6, 6)
        name_layout.addWidget(QtWidgets.QLabel("Nombre del Cliente:"))
        name_layout.addWidget(self.name)

        layout_items.addLayout(name_layout)

        nit_layout = QtWidgets.QVBoxLayout()
        nit_layout.setSpacing(6)
        self.nit = QtWidgets.QLineEdit("" if self.index is None else str(self.basededatos.datos["cliente"]["datos"][index][1]))
        self.nit.textChanged.connect(self.Revisor)
        self.nit.setTextMargins(6, 6, 6, 6)
        nit_layout.addWidget(QtWidgets.QLabel("NIT del Cliente:"))
        nit_layout.addWidget(self.nit)

        layout_items.addLayout(nit_layout)

        address_layout = QtWidgets.QVBoxLayout()
        address_layout.setSpacing(6)
        self.address = QtWidgets.QLineEdit("" if self.index is None else str(self.basededatos.datos["cliente"]["datos"][index][2]))
        self.address.textChanged.connect(self.Revisor)
        self.address.setTextMargins(6, 6, 6, 6)
        address_layout.addWidget(QtWidgets.QLabel("Dirección del Cliente:"))
        address_layout.addWidget(self.address)

        layout_items.addLayout(address_layout)

        layout_buttons = QtWidgets.QHBoxLayout()
        layout_buttons.setContentsMargins(9, 0, 9, 9)

        cancel = QtWidgets.QPushButton("Cancelar ")
        cancel.setStyleSheet("""
        QPushButton
        {
            background: rgba(255, 0, 0, 100);
            color: white;
        }
        QPushButton:hover
        {	
            background: rgba(255, 0, 0, 120);
            color: white;
        }
        """)
        cancel.setIconSize(QtCore.QSize(31, 31))
        cancel.setIcon(QtGui.QIcon("icons/round_cancel_white_48dp.png"))
        cancel.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        cancel.clicked.connect(self.close)
        layout_buttons.addWidget(cancel)

        layout_buttons.addSpacerItem(
            QtWidgets.QSpacerItem(
                0, 10, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding
            )
        )

        self.save = QtWidgets.QPushButton("Guardar ")
        self.save.setStyleSheet("""
                QPushButton
                {
                    background: rgba(100, 255, 100, 80);
                    color: white;
                }
                QPushButton:hover
                {	
                    background: rgba(100, 255, 100, 100);
                    color: white;
                }
                """)
        self.save.setIconSize(QtCore.QSize(31, 31))
        self.save.setIcon(QtGui.QIcon("icons/round_check_circle_white_48dp.png"))
        self.save.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.save.clicked.connect(self.guardar_cliente)
        layout_buttons.addWidget(self.save)

        main_frame_layout.addLayout(layout_items)
        main_frame_layout.addLayout(layout_buttons)
        main_frame.setLayout(main_frame_layout)
        self.layout.addWidget(main_frame)
        self.setLayout(self.layout)
        self.Revisor()

    def Revisor(self):
        revisor = True

        if self.name.text() == "":
            revisor = False

        if self.nit.text() == "":
            revisor = False

        if self.address.text() == "":
            revisor = False

        if revisor:
            self.save.show()
        else:
            self.save.hide()

    def guardar_cliente(self):
        if self.index is None:
            self.basededatos.crear_cliente(name=self.name.text(), nit=self.nit.text(), address=self.address.text())
        else:
            self.basededatos.editar_cliente(
                index=self.index, name=self.name.text(), nit=self.nit.text(), address=self.address.text()
            )
        self.close()
        self.cerrado.emit()


# Ventana de cotización a través de correo electrónico
class VentanaCorreoCotizacion(QtWidgets.QDialog):
    def __init__(self, basededatos, index: int):
        super(VentanaCorreoCotizacion, self).__init__()
        self.setWindowFlags(QtCore.Qt.WindowFlags(QtCore.Qt.FramelessWindowHint))
        self.setAttribute(QtCore.Qt.WA_TranslucentBackground)
        self.layout = QtWidgets.QVBoxLayout()
        self.layout.setContentsMargins(0, 0, 0, 0)

        main_frame = QtWidgets.QFrame()
        main_frame.setObjectName("main_frame")
        main_frame_layout = QtWidgets.QVBoxLayout()
        main_frame_layout.setContentsMargins(0, 0, 0, 0)

        self.setStyleSheet("""
        *{
            border-radius: 10px;
            font-family:century-gothic;
            font-size:24px;
        }
        * QFrame {
            background-color: rgba(30, 50, 120, 255);
        }
        * QLabel{
            background-color: rgba(255, 255, 255, 0);
            color: rgba(255, 255, 255, 255);
            font-size: 14px;
        }
        * QPushButton
        {
            background: rgba(255, 200, 200, 80);
            color: white;
        }
        * QPushButton:hover
        {	
            background: rgba(255, 200, 200, 100);
            color: white;
        }
        * QPushButton:disabled {
            background-color: white;
            color: black;
        }
        * QScrollBar:horizontal {
            border: none;
            background: rgb(52, 59, 72);
            height: 8px;
            margin: 0px 21px 0 21px;
            border-radius: 0px;
        }
        * QScrollBar::handle:horizontal {
            background: rgb(189, 147, 249);
            min-width: 25px;
            border-radius: 4px
        }
        * QScrollBar::add-line:horizontal {
            border: none;
            background: rgb(55, 63, 77);
            width: 20px;
            border-top-right-radius: 4px;
            border-bottom-right-radius: 4px;
            subcontrol-position: right;
            subcontrol-origin: margin;
        }
        * QScrollBar::sub-line:horizontal {
            border: none;
            background: rgb(55, 63, 77);
            width: 20px;
            border-top-left-radius: 4px;
            border-bottom-left-radius: 4px;
            subcontrol-position: left;
            subcontrol-origin: margin;
        }
        * QScrollBar::up-arrow:horizontal, QScrollBar::down-arrow:horizontal
        {
             background: none;
        }
        * QScrollBar::add-page:horizontal, QScrollBar::sub-page:horizontal
        {
             background: none;
        }
        * QScrollBar:vertical {
            border: none;
            background: rgb(52, 59, 72);
            width: 8px;
            margin: 21px 0 21px 0;
            border-radius: 0px;
        }
        * QScrollBar::handle:vertical {	
           background: rgb(189, 147, 249);
           min-height: 25px;
           border-radius: 4px
        }
        * QScrollBar::add-line:vertical {
            border: none;
           background: rgb(55, 63, 77);
            height: 20px;
           border-bottom-left-radius: 4px;
           border-bottom-right-radius: 4px;
            subcontrol-position: bottom;
            subcontrol-origin: margin;
        }
        * QScrollBar::sub-line:vertical {
           border: none;
           background: rgb(55, 63, 77);
            height: 20px;
           border-top-left-radius: 4px;
           border-top-right-radius: 4px;
            subcontrol-position: top;
            subcontrol-origin: margin;
        }
        * QScrollBar::up-arrow:vertical, QScrollBar::down-arrow:vertical {
            background: none;
        }
        * QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical {
            background: none;
        }
        """)

        self.basededatos = basededatos
        self.index = index

        title = QtWidgets.QLabel("Cotización por Correo")
        title.setMargin(6)
        title.setAlignment(QtCore.Qt.AlignCenter)
        title.setStyleSheet("font-size:24px; background: rgba(0, 0, 0, 40)")
        main_frame_layout.addWidget(title)

        layout_items = QtWidgets.QHBoxLayout()
        layout_items.setContentsMargins(9, 0, 9, 0)
        layout_items.setSpacing(14)

        name_layout = QtWidgets.QVBoxLayout()
        name_layout.setSpacing(6)
        self.name = QtWidgets.QLineEdit()
        self.name.textChanged.connect(self.Revisor)
        self.name.setTextMargins(6, 6, 6, 6)
        name_layout.addWidget(QtWidgets.QLabel("Nombre del Cliente:"))
        name_layout.addWidget(self.name)

        layout_items.addLayout(name_layout)

        mail_layout = QtWidgets.QVBoxLayout()
        mail_layout.setSpacing(6)
        self.mail = QtWidgets.QLineEdit()
        self.mail.textChanged.connect(self.Revisor)
        self.mail.setTextMargins(6, 6, 6, 6)
        mail_layout.addWidget(QtWidgets.QLabel("Correo del Cliente:"))
        mail_layout.addWidget(self.mail)

        layout_items.addLayout(mail_layout)

        layout_buttons = QtWidgets.QHBoxLayout()
        layout_buttons.setContentsMargins(9, 0, 9, 9)

        cancel = QtWidgets.QPushButton("Cancelar ")
        cancel.setStyleSheet("""
        QPushButton
        {
            background: rgba(255, 0, 0, 100);
            color: white;
        }
        QPushButton:hover
        {	
            background: rgba(255, 0, 0, 120);
            color: white;
        }
        """)
        cancel.setIconSize(QtCore.QSize(31, 31))
        cancel.setIcon(QtGui.QIcon("icons/round_cancel_white_48dp.png"))
        cancel.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        cancel.clicked.connect(self.close)
        layout_buttons.addWidget(cancel)

        layout_buttons.addSpacerItem(
            QtWidgets.QSpacerItem(
                0, 10, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding
            )
        )

        self.save = QtWidgets.QPushButton("Enviar ")
        self.save.setStyleSheet("""
                QPushButton
                {
                    background: rgba(100, 255, 100, 80);
                    color: white;
                }
                QPushButton:hover
                {	
                    background: rgba(100, 255, 100, 100);
                    color: white;
                }
                """)
        self.save.setIconSize(QtCore.QSize(31, 31))
        self.save.setIcon(QtGui.QIcon("icons/round_check_circle_white_48dp.png"))
        self.save.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.save.clicked.connect(self.enviar_correo)
        layout_buttons.addWidget(self.save)

        main_frame_layout.addLayout(layout_items)
        main_frame_layout.addLayout(layout_buttons)
        main_frame.setLayout(main_frame_layout)
        self.layout.addWidget(main_frame)
        self.setLayout(self.layout)
        self.Revisor()

    def Revisor(self):
        revisor = True

        if self.name.text() == "":
            revisor = False

        if self.mail.text() == "":
            revisor = False

        if "@" not in self.mail.text():
            revisor = False

        if "." not in self.mail.text():
            revisor = False

        if revisor:
            self.save.show()
        else:
            self.save.hide()

    #Envío del mensaje por correo a través de un server S.M.T.P.
    def enviar_correo(self):
        info_producto = self.basededatos.datos["productos"]["datos"][self.index]
        smtp_server = "smtp.gmail.com"
        port = 587  # For starttls
        sender_email = "testermailpython1@gmail.com"
        password = "a1234567A"
        # Create a secure SSL context
        context = ssl.create_default_context()

        # Try to log in to server and send email
        try:
            server = smtplib.SMTP(smtp_server, port)
            server.ehlo()  # Can be omitted
            server.starttls(context=context)  # Secure the connection
            server.ehlo()  # Can be omitted
            server.login(sender_email, password)

            server.sendmail(
                sender_email,
                self.mail.text(),
                f"Subject: {'Cotizacion de Productos'}\n"
                f"Estimado {self.name.text()}\n\n"
                f"El precio de nuestro producto {info_producto[0]}\n\n"
                f"Es Q {float(info_producto[1]):,.2f}"
            )
        except Exception as e:
            # Print any error messages to stdout
            print(e)
        finally:
            server.quit()
        self.close()


# Ventana de creación de pedidos
class VentanaPedidos(QtWidgets.QDialog):
    cerrado = QtCore.pyqtSignal()

    def __init__(self, basededatos):
        super(VentanaPedidos, self).__init__()

        self.setMinimumWidth(900 - 300)
        self.setMinimumHeight(600 - 300)
        self.setMaximumHeight(400)

        self.setWindowFlags(QtCore.Qt.WindowFlags(QtCore.Qt.FramelessWindowHint))
        self.setAttribute(QtCore.Qt.WA_TranslucentBackground)
        self.layout = QtWidgets.QVBoxLayout()
        self.layout.setContentsMargins(0, 0, 0, 0)

        main_frame = QtWidgets.QFrame()
        main_frame.setObjectName("main_frame")
        main_frame_layout = QtWidgets.QVBoxLayout()
        main_frame_layout.setContentsMargins(0, 0, 0, 0)

        self.setStyleSheet("""
        *{
            border-radius: 10px;
            font-family:century-gothic;
            font-size:24px;
        }
        * QFrame {
            background-color: rgba(30, 50, 120, 255);
        }
        * QLabel{
            background-color: rgba(255, 255, 255, 0);
            color: rgba(255, 255, 255, 255);
        }
        * QPushButton
        {
            background: rgba(255, 200, 200, 80);
            color: white;
        }
        * QPushButton:hover
        {	
            background: rgba(255, 200, 200, 100);
            color: white;
        }
        * QPushButton:disabled {
            background-color: white;
            color: black;
        }
        * QScrollBar:horizontal {
            border: none;
            background: rgb(52, 59, 72);
            height: 8px;
            margin: 0px 21px 0 21px;
            border-radius: 0px;
        }
        * QScrollBar::handle:horizontal {
            background: rgb(189, 147, 249);
            min-width: 25px;
            border-radius: 4px
        }
        * QScrollBar::add-line:horizontal {
            border: none;
            background: rgb(55, 63, 77);
            width: 20px;
            border-top-right-radius: 4px;
            border-bottom-right-radius: 4px;
            subcontrol-position: right;
            subcontrol-origin: margin;
        }
        * QScrollBar::sub-line:horizontal {
            border: none;
            background: rgb(55, 63, 77);
            width: 20px;
            border-top-left-radius: 4px;
            border-bottom-left-radius: 4px;
            subcontrol-position: left;
            subcontrol-origin: margin;
        }
        * QScrollBar::up-arrow:horizontal, QScrollBar::down-arrow:horizontal
        {
             background: none;
        }
        * QScrollBar::add-page:horizontal, QScrollBar::sub-page:horizontal
        {
             background: none;
        }
        * QScrollBar:vertical {
            border: none;
            background: rgb(52, 59, 72);
            width: 8px;
            margin: 21px 0 21px 0;
            border-radius: 0px;
        }
        * QScrollBar::handle:vertical {	
           background: rgb(189, 147, 249);
           min-height: 25px;
           border-radius: 4px
        }
        * QScrollBar::add-line:vertical {
            border: none;
           background: rgb(55, 63, 77);
            height: 20px;
           border-bottom-left-radius: 4px;
           border-bottom-right-radius: 4px;
            subcontrol-position: bottom;
            subcontrol-origin: margin;
        }
        * QScrollBar::sub-line:vertical {
           border: none;
           background: rgb(55, 63, 77);
            height: 20px;
           border-top-left-radius: 4px;
           border-top-right-radius: 4px;
            subcontrol-position: top;
            subcontrol-origin: margin;
        }
        * QScrollBar::up-arrow:vertical, QScrollBar::down-arrow:vertical {
            background: none;
        }
        * QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical {
            background: none;
        }
        """)

        self.basededatos = basededatos

        self.client_index = None
        self.product_index = None

        title = QtWidgets.QLabel("Creación de Pedido")
        title.setMargin(6)
        title.setAlignment(QtCore.Qt.AlignCenter)
        title.setStyleSheet("font-size:24px; background: rgba(0, 0, 0, 40)")
        main_frame_layout.addWidget(title)

        layout_items = QtWidgets.QHBoxLayout()
        layout_items.setContentsMargins(9, 0, 9, 0)
        layout_items.setSpacing(14)

        self.paginas = PySlidingStackedWidget()

        widget_clientes = QtWidgets.QWidget()
        layout_clientes = QtWidgets.QVBoxLayout()
        layout_clientes.setContentsMargins(0, 0, 0, 0)

        title = QtWidgets.QLabel("Selección de Cliente")
        title.setMargin(6)
        title.setAlignment(QtCore.Qt.AlignCenter)
        title.setStyleSheet("font-size:24px; background: rgba(255, 255, 255, 20)")
        layout_clientes.addWidget(title)

        filter_layout = QtWidgets.QHBoxLayout()
        filter_layout.addWidget(QtWidgets.QLabel("Filtrar: "))
        self.filter_cliente = QtWidgets.QLineEdit()
        self.filter_cliente.textChanged.connect(self.obtener_clientes)
        filter_layout.addWidget(self.filter_cliente)
        layout_clientes.addLayout(filter_layout)

        widget = QtWidgets.QWidget()
        widget.setStyleSheet("""
                                    * {
                                        background-color: rgba(255, 255, 255, 0)
                                    }
                                    """)
        scrollArea = QtWidgets.QScrollArea()
        scrollArea.setStyleSheet("""
                                    * {
                                        background-color: rgba(255, 255, 255, 0)
                                    }
                                    """)
        scrollArea.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarAsNeeded)
        scrollArea.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarAsNeeded)
        scrollArea.setWidgetResizable(True)

        scrollLayout = QtWidgets.QVBoxLayout()
        scrollLayout.setContentsMargins(0, 0, 0, 0)
        scrollLayout.setSpacing(0)

        self.frameTable_clientes = QtWidgets.QWidget()
        self.frameTable_clientes.setStyleSheet("""
                                    * {
                                        background-color: rgba(255, 255, 255, 0)
                                    }
                                    """)
        layoutTable = QtWidgets.QVBoxLayout()
        layoutTable.setContentsMargins(0, 0, 9, 0)
        layoutTable.setSpacing(6)
        self.frameTable_clientes.setLayout(layoutTable)
        self.frameTable_clientes.setStyleSheet("""* QFrame{background: rgba(255, 255, 255, 0);}""")

        scrollLayout.addWidget(self.frameTable_clientes)

        scrollLayout.addSpacerItem(
            QtWidgets.QSpacerItem(
                0, 10, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding
            )
        )

        widget.setLayout(scrollLayout)
        scrollArea.setWidget(widget)
        layout_clientes.addWidget(scrollArea)

        widget_clientes.setLayout(layout_clientes)

        widget_productos = QtWidgets.QWidget()
        layout_productos = QtWidgets.QVBoxLayout()
        layout_productos.setContentsMargins(0, 0, 0, 0)

        title = QtWidgets.QLabel("Selección de Producto")
        title.setMargin(6)
        title.setAlignment(QtCore.Qt.AlignCenter)
        title.setStyleSheet("font-size:24px; background: rgba(255, 255, 255, 20)")
        layout_productos.addWidget(title)

        filter_layout = QtWidgets.QHBoxLayout()
        filter_layout.addWidget(QtWidgets.QLabel("Filtrar: "))
        self.filter_producto = QtWidgets.QLineEdit()
        self.filter_producto.textChanged.connect(self.obtener_productos)
        filter_layout.addWidget(self.filter_producto)
        layout_productos.addLayout(filter_layout)

        widget = QtWidgets.QWidget()
        widget.setStyleSheet("""
                                            * {
                                                background-color: rgba(255, 255, 255, 0)
                                            }
                                            """)
        scrollArea = QtWidgets.QScrollArea()
        scrollArea.setStyleSheet("""
                                            * {
                                                background-color: rgba(255, 255, 255, 0)
                                            }
                                            """)
        scrollArea.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarAsNeeded)
        scrollArea.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarAsNeeded)
        scrollArea.setWidgetResizable(True)

        scrollLayout = QtWidgets.QVBoxLayout()
        scrollLayout.setContentsMargins(0, 0, 0, 0)
        scrollLayout.setSpacing(0)

        self.frameTable_productos = QtWidgets.QWidget()
        self.frameTable_productos.setStyleSheet("""
                                            * {
                                                background-color: rgba(255, 255, 255, 0)
                                            }
                                            """)
        layoutTable = QtWidgets.QVBoxLayout()
        layoutTable.setContentsMargins(9, 0, 9, 0)
        layoutTable.setSpacing(6)
        self.frameTable_productos.setLayout(layoutTable)
        self.frameTable_productos.setStyleSheet("""* QFrame{background: rgba(255, 255, 255, 0);}""")

        scrollLayout.addWidget(self.frameTable_productos)

        scrollLayout.addSpacerItem(
            QtWidgets.QSpacerItem(
                0, 10, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding
            )
        )

        widget.setLayout(scrollLayout)
        scrollArea.setWidget(widget)
        layout_productos.addWidget(scrollArea)

        widget_productos.setLayout(layout_productos)

        widget_cantidad = QtWidgets.QWidget()
        layout_cantidad = QtWidgets.QVBoxLayout()
        layout_cantidad.setContentsMargins(0, 0, 0, 0)

        layout_cantidad.addSpacerItem(
            QtWidgets.QSpacerItem(
                0, 10, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding
            )
        )

        title = QtWidgets.QLabel("Cantidad de Productos")
        title.setMargin(6)
        title.setAlignment(QtCore.Qt.AlignCenter)
        title.setStyleSheet("font-size:24px; background: rgba(255, 255, 255, 20)")
        layout_cantidad.addWidget(title)

        self.quantity = QtWidgets.QSpinBox()
        self.quantity.lineEdit().setTextMargins(6, 6, 6, 6)
        self.quantity.setAlignment(QtCore.Qt.AlignHCenter | QtCore.Qt.AlignVCenter)
        self.quantity.setButtonSymbols(QtWidgets.QAbstractSpinBox().NoButtons)
        self.quantity.setMinimum(0)
        self.quantity.setMaximum(999999999)
        self.quantity.valueChanged.connect(self.Revisor)
        self.quantity.clear()
        layout_cantidad.addWidget(self.quantity)

        layout_cantidad.addSpacerItem(
            QtWidgets.QSpacerItem(
                0, 10, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding
            )
        )

        widget_cantidad.setLayout(layout_cantidad)

        self.paginas.addWidget(widget_clientes)
        self.paginas.addWidget(widget_productos)
        self.paginas.addWidget(widget_cantidad)

        layout_items.addWidget(self.paginas)

        layout_buttons = QtWidgets.QHBoxLayout()
        layout_buttons.setContentsMargins(9, 0, 9, 9)

        cancel = QtWidgets.QPushButton("Cancelar ")
        cancel.setStyleSheet("""
        QPushButton
        {
            background: rgba(255, 0, 0, 100);
            color: white;
        }
        QPushButton:hover
        {	
            background: rgba(255, 0, 0, 120);
            color: white;
        }
        """)
        cancel.setIconSize(QtCore.QSize(31, 31))
        cancel.setIcon(QtGui.QIcon("icons/round_cancel_white_48dp.png"))
        cancel.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        cancel.clicked.connect(self.close)
        layout_buttons.addWidget(cancel)

        layout_buttons.addSpacerItem(
            QtWidgets.QSpacerItem(
                0, 10, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Fixed
            )
        )

        self.save = QtWidgets.QPushButton("Guardar ")
        self.save.setStyleSheet("""
                QPushButton
                {
                    background: rgba(100, 255, 100, 80);
                    color: white;
                }
                QPushButton:hover
                {	
                    background: rgba(100, 255, 100, 100);
                    color: white;
                }
                """)
        self.save.setIconSize(QtCore.QSize(31, 31))
        self.save.setIcon(QtGui.QIcon("icons/round_check_circle_white_48dp.png"))
        self.save.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.save.clicked.connect(self.finalizar_pedido)
        layout_buttons.addWidget(self.save)

        main_frame_layout.addLayout(layout_items)
        main_frame_layout.setSpacing(6)
        main_frame_layout.addLayout(layout_buttons)
        main_frame.setLayout(main_frame_layout)
        self.layout.addWidget(main_frame)
        self.setLayout(self.layout)
        self.paginas.currentChanged.connect(self.Revisor)
        self.Revisor()
        self.obtener_clientes()
        self.obtener_productos()

    def obtener_clientes(self):
        for _ in reversed(range(self.frameTable_clientes.layout().count())):
            self.frameTable_clientes.layout().itemAt(_).widget().setParent(None)

        for i, client in enumerate(self.basededatos.datos["cliente"]["datos"]):
            if self.filter_cliente.text() in f"{client[0]} {client[1]} {client[2]}":
                frame = AnimatedFrame()
                frame.setStyleSheet("""
                                    * QFrame{
                                        background-color: rgba(255, 255, 255, 0)
                                    }
                                    * QLabel{
                                        background-color: rgba(255, 255, 255, 40);
                                        color: rgba(255, 255, 255, 255)
                                    }
                                    * QPushButton
                                    {
                                        background: rgba(255, 200, 200, 80);
                                        color: white;
                                    }
                                    * QPushButton:hover
                                    {	
                                        background: rgba(255, 200, 200, 100);
                                        color: white;
                                    }
                                    """)
                frame.setMinimumHeight(31)
                frame_layout = QtWidgets.QHBoxLayout()
                frame_layout.setContentsMargins(0, 0, 0, 0)

                quantity = QtWidgets.QLabel(str(client[1]))
                quantity.setToolTip("NIT")
                quantity.setMinimumWidth(200)
                quantity.setAlignment(QtCore.Qt.AlignCenter)
                quantity.setSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Preferred)
                quantity.setMargin(6)
                frame_layout.addWidget(quantity)

                name = QtWidgets.QLabel(client[0])
                name.setWordWrap(True)
                name.setMargin(6)
                frame_layout.addWidget(name)

                edit_button = QtWidgets.QPushButton()
                edit_button.setSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Preferred)
                edit_button.setIconSize(QtCore.QSize(31, 31))
                edit_button.setToolTip("Usar Cliente")
                edit_button.setIcon(QtGui.QIcon("icons/round_forward_white_48dp.png"))
                edit_button.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
                edit_button.setAccessibleName(f"use_{i}")
                edit_button.clicked.connect(self.seleccionar_cliente)
                frame_layout.addWidget(edit_button)

                frame.setLayout(frame_layout)
                self.frameTable_clientes.layout().addWidget(frame)

    def obtener_productos(self):
        for _ in reversed(range(self.frameTable_productos.layout().count())):
            self.frameTable_productos.layout().itemAt(_).widget().setParent(None)

        for i, product in enumerate(self.basededatos.datos["productos"]["datos"]):
            product[1]
            if self.filter_producto.text() in f"{product[0]} {product[1]} {product[2]}" and product[1] > 0:
                frame = AnimatedFrame()
                frame.setStyleSheet("""
                                    * QFrame{
                                        background-color: rgba(255, 255, 255, 0)
                                    }
                                    * QLabel{
                                        background-color: rgba(255, 255, 255, 40);
                                        color: rgba(255, 255, 255, 255)
                                    }
                                    * QPushButton
                                    {
                                        background: rgba(255, 200, 200, 80);
                                        color: white;
                                    }
                                    * QPushButton:hover
                                    {	
                                        background: rgba(255, 200, 200, 100);
                                        color: white;
                                    }
                                    """)
                frame.setMinimumHeight(31)
                frame_layout = QtWidgets.QHBoxLayout()
                frame_layout.setContentsMargins(0, 0, 0, 0)

                quantity = QtWidgets.QLabel(f"{product[2]}".zfill(3))
                quantity.setToolTip("Existencias")
                quantity.setAlignment(QtCore.Qt.AlignHCenter)
                quantity.setSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Preferred)
                quantity.setMargin(6)
                frame_layout.addWidget(quantity)

                name = QtWidgets.QLabel(product[0])
                name.setWordWrap(True)
                name.setMargin(6)
                frame_layout.addWidget(name)

                price = QtWidgets.QLabel(f"Q {float(product[1]):,.2f}")
                price.setSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Preferred)
                price.setAlignment(QtCore.Qt.AlignRight)
                price.setToolTip("Precio")
                price.setMargin(6)
                price.setMinimumWidth(150)
                frame_layout.addWidget(price)

                edit_button = QtWidgets.QPushButton()
                edit_button.setSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Preferred)
                edit_button.setIconSize(QtCore.QSize(31, 31))
                edit_button.setToolTip("Usar Producto")
                edit_button.setIcon(QtGui.QIcon("icons/round_forward_white_48dp.png"))
                edit_button.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
                edit_button.setAccessibleName(f"useProd_{i}")
                edit_button.clicked.connect(self.seleccionar_producto)
                frame_layout.addWidget(edit_button)

                frame.setLayout(frame_layout)
                self.frameTable_productos.layout().addWidget(frame)

    def seleccionar_cliente(self):
        self.client_index = int(self.sender().accessibleName().split("_")[-1])
        self.paginas.slideInIdx(1)

    def seleccionar_producto(self):
        self.producto_index = int(self.sender().accessibleName().split("_")[-1])
        self.quantity.setMaximum(self.basededatos.datos["productos"]["datos"][self.producto_index][2])
        self.quantity.clear()
        self.paginas.slideInIdx(2)

    def Revisor(self):
        revisor = True

        if self.paginas.currentIndex() != 2:
            revisor = False

        if self.quantity.value() == 0:
            revisor = False

        if revisor:
            self.save.show()
        else:
            self.save.hide()

    def finalizar_pedido(self):
        name = self.basededatos.datos["cliente"]["datos"][self.client_index][0]
        product_info = self.basededatos.datos["productos"]["datos"][self.producto_index]
        product = product_info[0]
        quantity = self.quantity.value()
        price = quantity * product_info[1]

        self.basededatos.actualizar_existencia(index=self.producto_index, minus=quantity)
        self.basededatos.crear_pedido(name=name, product=product, quantity=quantity, price=price)
        self.close()
        self.cerrado.emit()


# Páginas de la ventana principal
class Pagina(QtWidgets.QWidget):
    def __init__(self):
        super(Pagina, self).__init__()
        self.layout = QtWidgets.QVBoxLayout()
        self.layout.setContentsMargins(0, 0, 0, 0)
        self.title = QtWidgets.QLabel()
        self.title.setAlignment(QtCore.Qt.AlignLeft)
        self.title.setStyleSheet("font-size: 24px")

        self.buttons_layout = QtWidgets.QHBoxLayout()
        self.buttons_layout.setContentsMargins(9, 9, 9, 0)
        self.layout.addLayout(self.buttons_layout)
        self.buttons_layout.addWidget(self.title)

        widget = QtWidgets.QWidget()
        widget.setStyleSheet("""
                            * {
                                background-color: rgba(255, 255, 255, 0)
                            }
                            """)
        scrollArea = QtWidgets.QScrollArea()
        scrollArea.setStyleSheet("""
                            * {
                                background-color: rgba(255, 255, 255, 0)
                            }
                            """)
        scrollArea.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarAsNeeded)
        scrollArea.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarAsNeeded)
        scrollArea.setWidgetResizable(True)

        scrollLayout = QtWidgets.QVBoxLayout()
        scrollLayout.setContentsMargins(0, 0, 0, 0)
        scrollLayout.setSpacing(0)

        self.frameTable = QtWidgets.QWidget()
        self.frameTable.setStyleSheet("""
                            * {
                                background-color: rgba(255, 255, 255, 0)
                            }
                            """)
        layoutTable = QtWidgets.QVBoxLayout()
        layoutTable.setContentsMargins(9, 0, 9, 0)
        layoutTable.setSpacing(6)
        self.frameTable.setLayout(layoutTable)
        self.frameTable.setStyleSheet("""* QFrame{background: rgba(255, 255, 255, 0);}""")

        scrollLayout.addWidget(self.frameTable)

        scrollLayout.addSpacerItem(
            QtWidgets.QSpacerItem(
                0, 10, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding
            )
        )

        widget.setLayout(scrollLayout)
        scrollArea.setWidget(widget)
        self.layout.addWidget(scrollArea)

        self.setLayout(self.layout)


# Página de productos, hereda de Pagina
class PaginaProductos(Pagina):
    def __init__(self, basededatos):
        super(PaginaProductos, self).__init__()

        self.basededatos = basededatos

        self.title.setText("Productos")

        self.add_button = QtWidgets.QPushButton("Agregar Producto")
        self.add_button.setIconSize(QtCore.QSize(31, 31))
        self.add_button.setIcon(QtGui.QIcon("icons/round_add_white_48dp.png"))
        self.add_button.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.add_button.clicked.connect(self.crear_producto)
        self.buttons_layout.addWidget(self.add_button, alignment=QtCore.Qt.AlignRight)

        QtCore.QTimer.singleShot(500, self.obtener_productos)

    def obtener_productos(self):
        for _ in reversed(range(self.frameTable.layout().count())):
            self.frameTable.layout().itemAt(_).widget().setParent(None)

        for i, product in enumerate(self.basededatos.datos["productos"]["datos"]):
            frame = AnimatedFrame()
            frame.setStyleSheet("""
                                * QFrame{
                                    background-color: rgba(255, 255, 255, 0)
                                }
                                * QLabel{
                                    background-color: rgba(255, 255, 255, 40);
                                    color: rgba(255, 255, 255, 255)
                                }
                                * QPushButton
                                {
                                    background: rgba(255, 200, 200, 80);
                                    color: white;
                                }
                                * QPushButton:hover
                                {	
                                    background: rgba(255, 200, 200, 100);
                                    color: white;
                                }
                                """)
            frame.setMinimumHeight(31)
            frame_layout = QtWidgets.QHBoxLayout()
            frame_layout.setContentsMargins(0, 0, 0, 0)

            delete_button = QtWidgets.QPushButton()
            delete_button.setSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Preferred)
            delete_button.setIconSize(QtCore.QSize(31, 31))
            delete_button.setStyleSheet("""
            QPushButton {
                background: rgba(255, 100, 100, 150);
                color: white;
            }
            QPushButton:hover {	
                background: rgba(255, 100, 100, 200);
                color: white;
            }
            """)
            delete_button.setToolTip("Borrar Producto")
            delete_button.setIcon(QtGui.QIcon("icons/round_clear_white_48dp.png"))
            delete_button.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
            delete_button.setAccessibleName(f"delete_{i}")
            delete_button.clicked.connect(self.borrar_producto)
            frame_layout.addWidget(delete_button)

            quantity = QtWidgets.QLabel(f"{product[2]}".zfill(3))
            quantity.setToolTip("Existencias")
            quantity.setAlignment(QtCore.Qt.AlignHCenter)
            quantity.setSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Preferred)
            quantity.setMargin(6)
            frame_layout.addWidget(quantity)

            name = QtWidgets.QLabel(product[0])
            name.setWordWrap(True)
            name.setMargin(6)
            frame_layout.addWidget(name)

            price = QtWidgets.QLabel(f"Q {float(product[1]):,.2f}")
            price.setSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Preferred)
            price.setAlignment(QtCore.Qt.AlignRight)
            price.setToolTip("Precio")
            price.setMargin(6)
            price.setMinimumWidth(150)
            frame_layout.addWidget(price)

            edit_button = QtWidgets.QPushButton()
            edit_button.setSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Preferred)
            edit_button.setIconSize(QtCore.QSize(31, 31))
            edit_button.setToolTip("Editar Producto")
            edit_button.setIcon(QtGui.QIcon("icons/round_edit_white_48dp.png"))
            edit_button.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
            edit_button.setAccessibleName(f"edit_{i}")
            edit_button.clicked.connect(self.editar_producto)
            frame_layout.addWidget(edit_button)

            cotiz_button = QtWidgets.QPushButton()
            cotiz_button.setSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Preferred)
            cotiz_button.setIconSize(QtCore.QSize(31, 31))
            cotiz_button.setToolTip("Enviar Cotización por Correo")
            cotiz_button.setIcon(QtGui.QIcon("icons/round_alternate_email_white_48dp.png"))
            cotiz_button.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
            cotiz_button.setAccessibleName(f"send_{i}")
            cotiz_button.clicked.connect(self.cotizar_por_correo)
            frame_layout.addWidget(cotiz_button)

            frame.setLayout(frame_layout)
            self.frameTable.layout().addWidget(frame)
            frame.show()

    def crear_producto(self):
        self.dialog = CreacionEdicionProductos(basededatos=self.basededatos)
        self.dialog.cerrado.connect(self.obtener_productos)
        self.dialog.exec_()

    def borrar_producto(self):
        sender = self.sender()
        sender.parent().remove()
        self.basededatos.borrar_item(int(sender.accessibleName().split("_")[-1]), "productos")
        QtCore.QTimer.singleShot(350, self.obtener_productos)

    def editar_producto(self):
        self.dialog = CreacionEdicionProductos(basededatos=self.basededatos, index=int(self.sender().accessibleName().split("_")[-1]))
        self.dialog.cerrado.connect(self.obtener_productos)
        self.dialog.exec_()

    def cotizar_por_correo(self):
        self.dialog = VentanaCorreoCotizacion(basededatos=self.basededatos, index=int(self.sender().accessibleName().split("_")[-1]))
        self.dialog.exec_()


# Página de clientes, hereda de Pagina
class PaginaClientes(Pagina):
    def __init__(self, basededatos):
        super(PaginaClientes, self).__init__()
        self.basededatos = basededatos

        self.title.setText("Clientes")
        self.add_button = QtWidgets.QPushButton("Agregar Cliente")
        self.add_button.setIconSize(QtCore.QSize(31, 31))
        self.add_button.setIcon(QtGui.QIcon("icons/round_add_white_48dp.png"))
        self.add_button.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.add_button.clicked.connect(self.crear_cliente)
        self.buttons_layout.addWidget(self.add_button, alignment=QtCore.Qt.AlignRight)

        QtCore.QTimer.singleShot(500, self.obtener_clientes)

    def obtener_clientes(self):
        for _ in reversed(range(self.frameTable.layout().count())):
            self.frameTable.layout().itemAt(_).widget().setParent(None)

        for i, client in enumerate(self.basededatos.datos["cliente"]["datos"]):
            frame = AnimatedFrame()
            frame.setStyleSheet("""
                                * QFrame{
                                    background-color: rgba(255, 255, 255, 0)
                                }
                                * QLabel{
                                    background-color: rgba(255, 255, 255, 40);
                                    color: rgba(255, 255, 255, 255)
                                }
                                * QPushButton
                                {
                                    background: rgba(255, 200, 200, 80);
                                    color: white;
                                }
                                * QPushButton:hover
                                {	
                                    background: rgba(255, 200, 200, 100);
                                    color: white;
                                }
                                """)
            frame.setMinimumHeight(31)
            frame_layout = QtWidgets.QHBoxLayout()
            frame_layout.setContentsMargins(0, 0, 0, 0)

            delete_button = QtWidgets.QPushButton()
            delete_button.setSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Preferred)
            delete_button.setIconSize(QtCore.QSize(31, 31))
            delete_button.setStyleSheet("""
            QPushButton {
                background: rgba(255, 100, 100, 150);
                color: white;
            }
            QPushButton:hover {	
                background: rgba(255, 100, 100, 200);
                color: white;
            }
            """)
            delete_button.setToolTip("Borrar Cliente")
            delete_button.setIcon(QtGui.QIcon("icons/round_clear_white_48dp.png"))
            delete_button.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
            delete_button.setAccessibleName(f"delete_{i}")
            delete_button.clicked.connect(self.borrar_cliente)
            frame_layout.addWidget(delete_button)

            quantity = QtWidgets.QLabel(str(client[1]))
            quantity.setToolTip("NIT")
            quantity.setMinimumWidth(200)
            quantity.setAlignment(QtCore.Qt.AlignCenter)
            quantity.setSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Preferred)
            quantity.setMargin(6)
            frame_layout.addWidget(quantity)

            name = QtWidgets.QLabel(client[0])
            name.setWordWrap(True)
            name.setMargin(6)
            frame_layout.addWidget(name)

            address = QtWidgets.QLabel(client[2])
            address.setWordWrap(True)
            address.setMargin(6)
            frame_layout.addWidget(address)

            edit_button = QtWidgets.QPushButton()
            edit_button.setSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Preferred)
            edit_button.setIconSize(QtCore.QSize(31, 31))
            edit_button.setToolTip("Editar Cliente")
            edit_button.setIcon(QtGui.QIcon("icons/round_edit_white_48dp.png"))
            edit_button.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
            edit_button.setAccessibleName(f"edit_{i}")
            edit_button.clicked.connect(self.editar_cliente)
            frame_layout.addWidget(edit_button)

            frame.setLayout(frame_layout)
            self.frameTable.layout().addWidget(frame)
            frame.show()

    def crear_cliente(self):
        self.dialog = CreacionEdicionClientes(basededatos=self.basededatos)
        self.dialog.cerrado.connect(self.obtener_clientes)
        self.dialog.exec_()

    def borrar_cliente(self):
        sender = self.sender()
        sender.parent().remove()
        self.basededatos.borrar_item(int(sender.accessibleName().split("_")[-1]), "cliente")
        QtCore.QTimer.singleShot(350, self.obtener_clientes)

    def editar_cliente(self):
        self.dialog = CreacionEdicionClientes(basededatos=self.basededatos, index=int(self.sender().accessibleName().split("_")[-1]))
        self.dialog.cerrado.connect(self.obtener_clientes)
        self.dialog.exec_()


# Página de pedidos, hereda de Pagina
class PaginaPedidos(Pagina):
    pedido_finalizado = QtCore.pyqtSignal()

    def __init__(self, basededatos):
        super(PaginaPedidos, self).__init__()
        self.basededatos = basededatos

        self.title.setText("Pedidos")

        self.add_button = QtWidgets.QPushButton("Agregar Pedido")
        self.add_button.setIconSize(QtCore.QSize(31, 31))
        self.add_button.setIcon(QtGui.QIcon("icons/round_add_white_48dp.png"))
        self.add_button.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.add_button.clicked.connect(self.crear_pedido)
        self.buttons_layout.addWidget(self.add_button, alignment=QtCore.Qt.AlignRight)

        QtCore.QTimer.singleShot(500, self.cargar_pedidos)

    def cargar_pedidos(self):
        for _ in reversed(range(self.frameTable.layout().count())):
            self.frameTable.layout().itemAt(_).widget().setParent(None)

        for i, order in enumerate(self.basededatos.datos["pedido"]["datos"]):
            frame = AnimatedFrame()
            frame.setStyleSheet("""
                    * QFrame{
                        background-color: rgba(255, 255, 255, 0)
                    }
                    * QLabel{
                        background-color: rgba(255, 255, 255, 40);
                        color: rgba(255, 255, 255, 255)
                    }
                    * QPushButton
                    {
                        background: rgba(255, 200, 200, 80);
                        color: white;
                    }
                    * QPushButton:hover
                    {	
                        background: rgba(255, 200, 200, 100);
                        color: white;
                    }
                    """)
            frame.setMinimumHeight(31)
            frame_layout = QtWidgets.QHBoxLayout()
            frame_layout.setContentsMargins(0, 0, 0, 0)

            delete_button = QtWidgets.QPushButton()
            delete_button.setSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Preferred)
            delete_button.setIconSize(QtCore.QSize(31, 31))
            delete_button.setStyleSheet("""
            QPushButton {
                background: rgba(255, 100, 100, 150);
                color: white;
            }
            QPushButton:hover {	
                background: rgba(255, 100, 100, 200);
                color: white;
            }
            """)
            delete_button.setToolTip("Borrar Cliente")
            delete_button.setIcon(QtGui.QIcon("icons/round_clear_white_48dp.png"))
            delete_button.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
            delete_button.setAccessibleName(f"delete_{i}")
            delete_button.clicked.connect(self.borrar_pedido)
            frame_layout.addWidget(delete_button)

            client_name = QtWidgets.QLabel(order[0])
            client_name.setWordWrap(True)
            client_name.setMargin(6)
            frame_layout.addWidget(client_name)

            quantity = QtWidgets.QLabel(f"{order[2]}".zfill(3))
            quantity.setToolTip("Cantidad")
            quantity.setAlignment(QtCore.Qt.AlignHCenter)
            quantity.setSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Preferred)
            quantity.setMargin(6)
            frame_layout.addWidget(quantity)

            product_name = QtWidgets.QLabel(order[1])
            product_name.setWordWrap(True)
            product_name.setMargin(6)
            frame_layout.addWidget(product_name)

            order_price = QtWidgets.QLabel(f"Q {float(order[3]):,.2f}")
            order_price.setSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Preferred)
            order_price.setAlignment(QtCore.Qt.AlignRight)
            order_price.setToolTip("Valor Pedido")
            order_price.setMargin(6)
            order_price.setMinimumWidth(150)
            frame_layout.addWidget(order_price)

            frame.setLayout(frame_layout)
            self.frameTable.layout().addWidget(frame)
            frame.show()

    def crear_pedido(self):
        self.dialog = VentanaPedidos(basededatos=self.basededatos)
        self.dialog.cerrado.connect(self.finalizar_pedido)
        self.dialog.exec_()

    def finalizar_pedido(self):
        self.cargar_pedidos()
        self.pedido_finalizado.emit()

    def borrar_pedido(self):
        sender = self.sender()
        sender.parent().remove()
        self.basededatos.borrar_item(int(sender.accessibleName().split("_")[-1]), "pedido")
        QtCore.QTimer.singleShot(350, self.cargar_pedidos)


# Página de reportes, hereda de Pagina
class PaginaReportes(Pagina):
    def __init__(self, basededatos):
        super(PaginaReportes, self).__init__()
        self.title.setText("Informes")

        self.basededatos = basededatos

        self.boton_informe_cliente = QtWidgets.QPushButton("  Total de ventas por cliente  ")
        self.boton_informe_cliente.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.boton_informe_cliente.clicked.connect(self.reporte_cliente)
        self.buttons_layout.addWidget(self.boton_informe_cliente)

        self.boton_informe_producto = QtWidgets.QPushButton("  Total de ventas por producto  ")
        self.boton_informe_producto.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.boton_informe_producto.clicked.connect(self.reporte_producto)
        self.buttons_layout.addWidget(self.boton_informe_producto)

    def reporte_cliente(self):
        for _ in reversed(range(self.frameTable.layout().count())):
            self.frameTable.layout().itemAt(_).widget().setParent(None)

        result = {}
        for order in self.basededatos.datos["pedido"]["datos"]:
            if order[0] in list(result.keys()):
                result[order[0]]["cantidad"] += int(order[2])
                result[order[0]]["valor"] += float(order[3])
            else:
                result[order[0]] = {
                    "cantidad": int(order[2]),
                    "valor": float(order[3])
                }

        for order in result:
            frame = AnimatedFrame()
            frame.setStyleSheet("""
                                * QFrame{
                                    background-color: rgba(255, 255, 255, 0)
                                }
                                * QLabel{
                                    background-color: rgba(255, 255, 255, 40);
                                    color: rgba(255, 255, 255, 255)
                                }
                                * QPushButton
                                {
                                    background: rgba(255, 200, 200, 80);
                                    color: white;
                                }
                                * QPushButton:hover
                                {	
                                    background: rgba(255, 200, 200, 100);
                                    color: white;
                                }
                                """)
            frame.setMinimumHeight(31)
            frame_layout = QtWidgets.QHBoxLayout()
            frame_layout.setContentsMargins(0, 0, 0, 0)

            name = QtWidgets.QLabel(order)
            name.setWordWrap(True)
            name.setMargin(6)
            frame_layout.addWidget(name)

            quantity = QtWidgets.QLabel(str(result[order]['cantidad']).zfill(3))
            quantity.setSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Preferred)
            quantity.setWordWrap(True)
            quantity.setToolTip("Cantidad de Productos")
            quantity.setMargin(6)
            frame_layout.addWidget(quantity)

            price = QtWidgets.QLabel(f"Q {float(result[order]['valor']):,.2f}")
            price.setSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Preferred)
            price.setAlignment(QtCore.Qt.AlignRight)
            price.setToolTip("Precio")
            price.setMargin(6)
            price.setMinimumWidth(150)
            frame_layout.addWidget(price)

            frame.setLayout(frame_layout)
            self.frameTable.layout().addWidget(frame)
            frame.show()

    def reporte_producto(self):
        for _ in reversed(range(self.frameTable.layout().count())):
            self.frameTable.layout().itemAt(_).widget().setParent(None)

        result = {}
        for order in self.basededatos.datos["pedido"]["datos"]:
            if order[0] in list(result.keys()):
                result[order[1]]["cantidad"] += int(order[2])
                result[order[1]]["valor"] += float(order[3])
            else:
                result[order[1]] = {
                    "cantidad": int(order[2]),
                    "valor": float(order[3])
                }

        for order in result:
            frame = AnimatedFrame()
            frame.setStyleSheet("""
                                * QFrame{
                                    background-color: rgba(255, 255, 255, 0)
                                }
                                * QLabel{
                                    background-color: rgba(255, 255, 255, 40);
                                    color: rgba(255, 255, 255, 255)
                                }
                                * QPushButton
                                {
                                    background: rgba(255, 200, 200, 80);
                                    color: white;
                                }
                                * QPushButton:hover
                                {	
                                    background: rgba(255, 200, 200, 100);
                                    color: white;
                                }
                                """)
            frame.setMinimumHeight(31)
            frame_layout = QtWidgets.QHBoxLayout()
            frame_layout.setContentsMargins(0, 0, 0, 0)

            name = QtWidgets.QLabel(order)
            name.setWordWrap(True)
            name.setMargin(6)
            frame_layout.addWidget(name)

            quantity = QtWidgets.QLabel(str(result[order]['cantidad']).zfill(3))
            quantity.setSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Preferred)
            quantity.setWordWrap(True)
            quantity.setToolTip("Cantidad de Productos")
            quantity.setMargin(6)
            frame_layout.addWidget(quantity)

            price = QtWidgets.QLabel(f"Q {float(result[order]['valor']):,.2f}")
            price.setSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Preferred)
            price.setAlignment(QtCore.Qt.AlignRight)
            price.setToolTip("Precio")
            price.setMargin(6)
            price.setMinimumWidth(150)
            frame_layout.addWidget(price)

            frame.setLayout(frame_layout)
            self.frameTable.layout().addWidget(frame)
            frame.show()


# Página varios, hereda de Pagina
class PaginaVarios(Pagina):
    def __init__(self, basededatos):
        super(PaginaVarios, self).__init__()
        self.title.setText("Varios")
        self.basededatos = basededatos

        self.boton_seguridad = QtWidgets.QPushButton("Crear copia de seguridad de datos")
        self.boton_seguridad.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.boton_seguridad.clicked.connect(self.copia_seguridad)
        self.buttons_layout.addWidget(self.boton_seguridad)

    def copia_seguridad(self):
        server = "smtp.gmail.com"
        port = 587  # For starttls
        sender_email = "testermailpython1@gmail.com"
        password = "a1234567A"
        correo_coordinador = "dgonzalezp15@miumg.edu.gt"

        msg = MIMEMultipart()
        msg['From'] = sender_email
        msg['To'] = correo_coordinador
        msg['Date'] = formatdate(localtime=True)
        msg['Subject'] = 'Copia de Seguridad'
        msg.attach(MIMEText("Copia de Seguridad"))

        part = MIMEBase('application', "octet-stream")
        part.set_payload(open("Inventario.xlsx", "rb").read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', 'attachment; filename="WorkBook3.xlsx"')
        msg.attach(part)

        # context = ssl.SSLContext(ssl.PROTOCOL_SSLv3)
        # SSL connection only working on Python 3+
        smtp = smtplib.SMTP(server, port)
        smtp.starttls()
        smtp.login(sender_email, password)
        smtp.sendmail(sender_email, correo_coordinador, msg.as_string())
        smtp.quit()


# Ventana Principal
class MainWindow(QtWidgets.QMainWindow):
    def __init__(self):
        super(MainWindow, self).__init__()
        self.basededatos = BaseDatos()

        self.setMinimumWidth(900)
        self.setMinimumHeight(600)

        self.setStyleSheet("""
        *{
            border-radius: 10px;
            font-family:century-gothic;
            font-size:24px;
        }
        QMainWindow {
            background-color: rgba(30, 50, 120, 255)
        }
        * QFrame{
            background-color: rgba(255, 255, 255, 40)
        }
        * QLabel{
            background-color: rgba(255, 255, 255, 0);
            color: rgba(255, 255, 255, 255)
        }
        * QPushButton
        {
            background: rgba(255, 200, 200, 80);
            color: white;
        }
        * QPushButton:hover
        {	
            background: rgba(255, 200, 200, 100);
            color: white;
        }
        * QPushButton:disabled {
            background-color: white;
            color: rgba(30, 50, 120, 255);
        }
        * QScrollBar:horizontal {
            border: none;
            background: rgb(52, 59, 72);
            height: 8px;
            margin: 0px 21px 0 21px;
            border-radius: 0px;
        }
        * QScrollBar::handle:horizontal {
            background: rgb(189, 147, 249);
            min-width: 25px;
            border-radius: 4px
        }
        * QScrollBar::add-line:horizontal {
            border: none;
            background: rgb(55, 63, 77);
            width: 20px;
            border-top-right-radius: 4px;
            border-bottom-right-radius: 4px;
            subcontrol-position: right;
            subcontrol-origin: margin;
        }
        * QScrollBar::sub-line:horizontal {
            border: none;
            background: rgb(55, 63, 77);
            width: 20px;
            border-top-left-radius: 4px;
            border-bottom-left-radius: 4px;
            subcontrol-position: left;
            subcontrol-origin: margin;
        }
        * QScrollBar::up-arrow:horizontal, QScrollBar::down-arrow:horizontal
        {
             background: none;
        }
        * QScrollBar::add-page:horizontal, QScrollBar::sub-page:horizontal
        {
             background: none;
        }
        * QScrollBar:vertical {
            border: none;
            background: rgb(52, 59, 72);
            width: 8px;
            margin: 21px 0 21px 0;
            border-radius: 0px;
        }
        * QScrollBar::handle:vertical {	
           background: rgb(189, 147, 249);
           min-height: 25px;
           border-radius: 4px
        }
        * QScrollBar::add-line:vertical {
            border: none;
           background: rgb(55, 63, 77);
            height: 20px;
           border-bottom-left-radius: 4px;
           border-bottom-right-radius: 4px;
            subcontrol-position: bottom;
            subcontrol-origin: margin;
        }
        * QScrollBar::sub-line:vertical {
           border: none;
           background: rgb(55, 63, 77);
            height: 20px;
           border-top-left-radius: 4px;
           border-top-right-radius: 4px;
            subcontrol-position: top;
            subcontrol-origin: margin;
        }
        * QScrollBar::up-arrow:vertical, QScrollBar::down-arrow:vertical {
            background: none;
        }
        * QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical {
            background: none;
        }
        """)

        self.central_widget = QtWidgets.QWidget()
        self.layout = QtWidgets.QVBoxLayout()

        self.paginas = PySlidingStackedWidget()
        self.pagina_productos = PaginaProductos(self.basededatos)
        self.paginas.addWidget(self.pagina_productos)
        self.pagina_clientes = PaginaClientes(self.basededatos)
        self.paginas.addWidget(self.pagina_clientes)
        self.pagina_pedidos = PaginaPedidos(self.basededatos)
        self.pagina_pedidos.pedido_finalizado.connect(self.orden_finalizada)
        self.paginas.addWidget(self.pagina_pedidos)
        self.pagina_reportes = PaginaReportes(self.basededatos)
        self.paginas.addWidget(self.pagina_reportes)
        self.pagina_varios = PaginaVarios(self.basededatos)
        self.paginas.addWidget(self.pagina_varios)
        self.paginas.currentChanged[int].connect(self.cambio_pagina)

        buttons_layout = QtWidgets.QHBoxLayout()
        buttons_layout.setContentsMargins(0, 0, 0, 0)

        self.productos_boton = QtWidgets.QPushButton("Productos")
        self.productos_boton.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.productos_boton.clicked.connect(lambda: self.paginas.slideInIdx(0))
        buttons_layout.addWidget(self.productos_boton)

        self.clientes_boton = QtWidgets.QPushButton("Clientes")
        self.clientes_boton.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.clientes_boton.clicked.connect(lambda: self.paginas.slideInIdx(1))
        buttons_layout.addWidget(self.clientes_boton)

        self.pedidos_boton = QtWidgets.QPushButton("Pedidos")
        self.pedidos_boton.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.pedidos_boton.clicked.connect(lambda: self.paginas.slideInIdx(2))
        buttons_layout.addWidget(self.pedidos_boton)

        self.informes_boton = QtWidgets.QPushButton("Informes")
        self.informes_boton.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.informes_boton.clicked.connect(lambda: self.paginas.slideInIdx(3))
        buttons_layout.addWidget(self.informes_boton)

        self.varios_boton = QtWidgets.QPushButton("Varios")
        self.varios_boton.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.varios_boton.clicked.connect(lambda: self.paginas.slideInIdx(4))
        buttons_layout.addWidget(self.varios_boton)

        self.productos_boton.setMinimumHeight(20)
        self.clientes_boton.setMinimumHeight(20)
        self.pedidos_boton.setMinimumHeight(20)
        self.informes_boton.setMinimumHeight(20)
        self.varios_boton.setMinimumHeight(20)

        self.layout.addLayout(buttons_layout)
        self.layout.addWidget(self.paginas)

        self.central_widget.setLayout(self.layout)
        self.setCentralWidget(self.central_widget)
        self.cambio_pagina(0)

    def cambio_pagina(self, index):
        self.productos_boton.setEnabled(True)
        self.clientes_boton.setEnabled(True)
        self.pedidos_boton.setEnabled(True)
        self.informes_boton.setEnabled(True)
        self.varios_boton.setEnabled(True)

        if index == 0:
            self.productos_boton.setEnabled(False)
        elif index == 1:
            self.clientes_boton.setEnabled(False)
        elif index == 2:
            self.pedidos_boton.setEnabled(False)
        elif index == 3:
            self.informes_boton.setEnabled(False)
        elif index == 4:
            self.varios_boton.setEnabled(False)

    def orden_finalizada(self):
        self.pagina_productos.obtener_productos()
        self.pagina_clientes.obtener_clientes()


def except_hook(cls, exception, traceback):
    sys.__excepthook__(cls, exception, traceback)


# Inicio de app
myappid = u'mycompany.myproduct.subproduct.version'
ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(myappid)
if __name__ == "__main__":
    sys.excepthook = except_hook
    app = QtWidgets.QApplication(sys.argv)
    app.setWindowIcon(QtGui.QIcon("icon.png"))
    app.setApplicationName("Control de Inventario")
    app.setApplicationDisplayName("Control de Inventario")
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())
